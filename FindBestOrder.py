import random
import openpyxl
import itertools


def Finddata(dataset, test, k=2):  # 找出基于元特征相似度最高的k个数据集
    readbook = openpyxl.load_workbook('meta_feature_similarity.xlsx')
    sheet = readbook.get_sheet_by_name('Sheet1')
    lenth = len(dataset)
    d = {}
    for i in range(2, lenth + 3):
        if sheet.cell(i, 1).value == test:
            for j in range(2, lenth + 3):
                if sheet.cell(1, j).value != test:
                    d[float(sheet.cell(i, j).value)] = j
    similarity = list(d.keys())
    similarity.sort(reverse=True)
    result = []
    result_data = {}
    for i in range(k):
        result.append(d[similarity[i]])
        result_data[str(sheet.cell(1, int(d[similarity[i]])).value)] = similarity[i]
    return result_data


def main():
    test = "Fodors_Zagats"
    k = 2
    dataset = ["DBLP_ACM", "DBLP_GoogleScholar", "Amazon_Google", "Walmart_Amazon", "Beer"]
    result_data = Finddata(dataset, test, k)
    dataset = list(result_data.keys())
    d = {}
    old_best_order = {}
    for i in range(len(dataset)):
        d[dataset[i]] = {}
        old_best_order[dataset[i]] = {}
    attributes = {}
    for data in dataset:
        attributes[data] = []
        with open("data/er_magellan/Structured/" + data + "/test.txt", "r") as f:
            i = f.readline()
            i = i.split("\t")
            i = i[0].split("COL")
            size = len(i)
            for j in range(1, size):
                val = i[j].split("VAL")
                attributes[data].append(val[0])
        data1 = data + ".txt_"
        data2 = test + ".txt"
        if data > test:
            data2 = data + ".txt"
            data1 = test + ".txt_"
        with open("Similarity/"+data1 + data2, "r") as f:
            for i in f.readlines():
                i = i.split(":")
                j = i[0].split(" and ")
                another = j[1].split(" embeddings")
                similarity = i[1].split("\n")
                similarity = similarity[0].split(" ")
                if d[data].get(j[0]) is None:
                    d[data][j[0]] = {}
                d[data][j[0]][another[0]] = float(similarity[1])
        readbook = openpyxl.load_workbook("best_order/" + data + ".xlsx")
        sheet = readbook.get_sheet_by_name('Sheet1')
        cols = sheet.max_column
        for i in range(cols):
            old_best_order[data][sheet.cell(1, i+1).value] = int(sheet.cell(2, i+1).value)
    print(old_best_order)  # 基于元特征相似度最高的k个数据集的最优属性序
    print(d)  # 目标数据集与相似度最高的k个数据集的属性相似度
    file = "data/er_magellan/Structured/"+test+"/test.txt"
    a = []
    with open(file, "r") as f:
        i = f.readline()
        i = i.split("\t")
        i = i[0].split("COL")
        size = len(i)
        for j in range(1, size):
            val = i[j].split("VAL")
            a.append(val[0])
    print(a)  # 目标数据集的属性
    a_match = {}
    for data in dataset:
        a_match[data] = {}
        T = a.copy()
        T.extend(attributes[data])
        print(T)
        while len(T) != 0:
            for attribute in T:
                if len(d[data][attribute]) != 0:
                    dict = d[data][attribute]
                    closest = ""
                    closest_distance = 0
                    for key in dict.keys():
                        if dict[key] > closest_distance:
                            closest = key
                            closest_distance = dict[key]
                    closest2 = ""
                    closest_distance2 = 0
                    dict2 = d[data][closest]
                    for key in dict2.keys():
                        if dict2[key] > closest_distance2:
                            closest2 = key
                            closest_distance2 = dict2[key]
                    if attribute == closest2:
                        if attribute in a:
                            a_match[data][attribute] = closest
                        else:
                            a_match[data][closest] = attribute
                        T.remove(attribute)
                        T.remove(closest)
                    else:
                        del d[data][attribute][closest]
                        del d[data][closest][attribute]
                else:
                    T.remove(attribute)
    print(a_match)
    print(result_data)
    order = []
    for i in range(len(a)):
        order.append(i+1)
    orders = list(itertools.permutations(order, len(a)))
    new_best_order = []
    maxQ = float('-inf')
    for i in range(len(orders)):
        Q = 0
        order = orders[i]
        for data in dataset:
            lenth = len(a_match[data])
            attribute_list = list(a_match[data].keys())
            for j in range(lenth-1):
                for k in range(j+1, lenth):
                    j_order = order[a.index(attribute_list[j])]
                    k_order = order[a.index(attribute_list[k])]
                    j_align = old_best_order[data][a_match[data][attribute_list[j]]]
                    k_align = old_best_order[data][a_match[data][attribute_list[k]]]
                    value = result_data[data] * d[data][attribute_list[j]][a_match[data][attribute_list[j]]] * d[data][attribute_list[k]][a_match[data][attribute_list[k]]]
                    if (j_order-k_order)*(j_align-k_align) < 0:
                        value = -value
                    Q = Q + value
        print(Q)
        if Q > maxQ:
            maxQ = Q
            new_best_order = orders[i]
            print(new_best_order)
    print("Best order:", new_best_order)
    print("Best Q:", maxQ)


if __name__ == "__main__":
    main()
