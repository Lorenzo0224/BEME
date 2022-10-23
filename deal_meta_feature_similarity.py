import math
import openpyxl

dataset = ["Beer", "Amazon_Google", "DBLP_ACM", "DBLP_GoogleScholar", "Walmart_Amazon", "Fodors_Zagats"]
data = [[1, 1, 0.25, 4, 12, 4, 1.6500, 0.4167, 0.0833, 6.3167, 3.8447],
        # [1, 1, 0.25, 4, 182, 14, 2.3720, 0.4505, 0.0055, 7.2873, 3.99],
        [1, 1, 0.33, 3, 274, 50, 4.4388, 0.1892, 0.0068, 605.7034, 16737442.6866],
        # [1, 1, 0.33, 3, 4586, 221, 5.3348, 0.1818, 0.0004, 464.7625, 7870391.8064],
        [1, 1, 0.25, 4, 4946, 10, 3.1128, 0.1723, 0.0275, 1998.3421, 8.2918],
        [1, 1, 0.25, 4, 11484, 1163, 5.4275, 0.1615, 0, 1996.5537, 1892.3488],
        [1, 2, 0.2, 5, 4098, 249, 5.4601, 0.2123, 0.0002, 788.4750, 5640852.6582],
        [1, 2, 0.1667, 6, 378, 21, 2.9875, 0.2937, 0.0026, 290.1852, 64258.2408]]
data_min = []
data_max = []
for i in range(11):
    data_max.append(data[0][i])
    data_min.append(data[0][i])
for i in range(1, 6):
    for j in range(11):
        if data[i][j] > data_max[j]:
            data_max[j] = data[i][j]
        if data[i][j] < data_min[j]:
            data_min[j] = data[i][j]
for i in range(6):
    for j in range(11):
        if data_max[j] == data_min[j]:
            data[i][j] = 0
        else:
            data[i][j] = (data[i][j] - data_min[j]) / (data_max[j] - data_min[j])
similarity = []
b = []
for i in range(6):
    r = 0
    for j in range(11):
        r += data[i][j] * data[i][j]
    b.append(r)
for i in range(6):
    result = []
    for j in range(6):
        a = 0
        for k in range(11):
            a = a + data[i][k] * data[j][k]
        result.append(a / (math.sqrt(b[i]) * math.sqrt(b[j])))
    similarity.append(result)
print(similarity)
outwb = openpyxl.Workbook()
outws = outwb.create_sheet(index=0)  # 创建sheet
for i in range(2, 8):
    outws.cell(1, i).value = dataset[i - 2]
    outws.cell(i, 1).value = dataset[i - 2]
for row in range(2, 8):
    for col in range(2, 8):
        outws.cell(row, col).value = similarity[row - 2][col - 2]  # 写文件
saveExcel = "meta_feature_similarity.xlsx"
outwb.save(saveExcel)
